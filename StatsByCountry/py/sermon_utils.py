
#!/usr/bin/env python
# Utility Functions to Initialize Simulator Cells.
# Dave Babbitt <dave.babbitt@gmail.com>
# Author: Dave Babbitt, Data Scientist
# coding: utf-8
"""
SermonScrapingUtilities: A set of utility functions common to sermon audio web scraping
"""
from bs4 import BeautifulSoup as bs
from datetime import datetime
from datetime import timedelta
from dateutil import relativedelta
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from pathlib import Path
from pydub import AudioSegment
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from shutil import copyfile
from urllib.request import urlretrieve
import math
import numpy as np
import os
import pandas as pd
import random
import re
import storage as s
import time
import urllib
import warnings
warnings.filterwarnings("ignore")

class SermonScrapingUtilities(object):
    """This class implements the core of the utility functions
    needed to scrape the sermon mp3s.
    
    Examples
    --------
    
    >>> import sermon_utils
    >>> u = sermon_utils.SermonScrapingUtilities()
    """
    
    def __init__(self):
        self.s = s.Storage()
        
        # Obscuration error pattern
        self.obscure_regex = re.compile('<([^ ]+)[^>]*class="([^"]+)"[^>]*>')
        
        # Sermon Archive URL
        self.squarespace_url = 'https://sphere-green-kslg.squarespace.com/config/pages/5ca546e7419202e545ca8977'
        
        # Sermonaudio.com Submit URL
        self.submitsermon_url = 'https://www.sermonaudio.com/submitsermon.asp'
        
        # Sermonaudio.com signin URL
        self.signin_url = 'https://www.sermonaudio.com/signin.asp'
        
        # Sermonaudio.com search formatted string
        self.search_formatted_str = 'https://www.sermonaudio.com/search.asp?keyword={}&SpeakerOnly=true&keywordwithin={}'
        
        # Sermonaudio.com general formatted string
        self.formatted_str = 'https://www.sermonaudio.com/{}'
        
        # Full xpath for the edit button
        div_class_list = 'sqs-damask'.split(' ')
        contains_list = ["contains(@class, '{}')".format(c) for c in div_class_list]
        xpath_list = ["html", "body", "div[{}]".format(' and '.join(contains_list)),
                      "div", "div[1]", "div", "div", "div", "div", "div[2]", "div[1]",
                      "div[3]", "div", "div", "div", "div", "div[3]", "div[1]"]
        self.button_xpath = "/{}".format('/'.join(xpath_list))
        
        # Full xpath for the search record
        span_class_list = ['title']
        span_contains_list = ["contains(@class, '{}')".format(c) for c in span_class_list]
        xpath_list = ["span[{}]".format(' and '.join(span_contains_list))]
        self.title_xpath = "//{}".format('/'.join(xpath_list))
        
        # Full xpath for the search field
        div_class_list = ['search-filter-input']
        div_contains_list = ["contains(@class, '{}')".format(c) for c in div_class_list]
        input_class_list = ['search-filter']
        input_contains_list = ["contains(@class, '{}')".format(c) for c in input_class_list]
        xpath_list = ["div[{}]".format(' and '.join(div_contains_list)),
                      "input[{}]".format(' and '.join(input_contains_list))]
        self.input_xpath = "//{}".format('/'.join(xpath_list))
        
        # CSS selector for the login button
        #self.login_css = 'button [data-test="login-button"]'
        self.login_css = '#renderTarget > div > div:nth-child(2) > div > div:nth-child(1) > div > div:nth-child(2) > div:nth-child(1) > div > button'
        self.login_xpath = '/html/body/div[1]/div/div[2]/div/div[1]/div/div[1]/div[1]/div/button'
        
        # Xpath list for audio processing div
        div_class_str = 'flyout dialog-editor-block-audio yui3-dd-draggable visible'
        self.audio_xpath_list = ['html', 'body',
                                 self.get_div(div_class_str),
                                 self.get_div('main-container'),
                                 self.get_div('body-block scrollable'),
                                 self.get_div('dialog-tab-item'),
                                 self.get_div('sqs-multi-frame data-state-loaded name-file-location-frame variable-height'),
                                 self.get_div('sqs-multi-frame-content yui3-widget-content-expanded'),
                                 self.get_div('multi-frame-wrapper animate-switch'),
                                 self.get_div('frame'),
                                 self.get_div('yui3-widget sqs-widget sqs-data-widget sqs-dialog-field sqs-file yui3-filerequiresprocessing sqs-file-audio-processing data-state-loaded name-audioAsset'),
                                 self.get_div('sqs-file-audio-processing-content yui3-widget-content-expanded')]
        
        # Full xpath for the progress bar
        bar_xpath = '/html/body/div/div[1]/div/div[1]/div'
        xpath_list = self.audio_xpath_list + [
                                              self.get_div('progress-container'),
                                              self.get_div('progress'),
                                              self.get_div('bar'),
                                              self.get_div('bar-inner')
                                              ]
        self.bar_xpath = "/{}".format('/'.join(xpath_list))
        
        # Full xpath for the spinner
        spinner_xpath = '/html/body/div/div[2]/div/div[1]/div/div'
        xpath_list = self.audio_xpath_list + [
                                              self.get_div('progress-container'),
                                              self.get_div('progress'),
                                              self.get_div('spinner'),
                                              self.get_div('yui3-widget sqs-spin dark'),
                                              self.get_div('sqs-spin-content yui3-widget-content-expanded')
                                              ]
        self.spinner_xpath = "/{}".format('/'.join(xpath_list))
        
        # Full xpath for the file preview div
        preview_css = 'span.file-meta.file-meta-name'
        preview_xpath = '/html/body/div[27]/div/div[2]/div[1]/div[3]/div/div/div[1]/div/div/div[4]/span[class="file-meta file-meta-name"]'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui flyout light buttons-right dialog-editor-block-audio yui3-dd-draggable visible'
        xpath_list = ['html', 'body', self.get_div(div_class_str), 'div', self.get_div('body-block scrollable'), self.get_div('tab-wrapper dialog-tab-item'),
                      self.get_div('yui3-widget sqs-widget sqs-data-widget sqs-dialog-field sqs-multi-frame data-state-loaded name-file-location-frame variable-height'),
                      'div', 'div', 'div[data-name="embedded"]', 'div',
                      'div', 'div[class="sqs-file-preview"]', 'span[class="file-meta file-meta-name"]']
        self.preview_xpath = "/{}".format('/'.join(xpath_list[:-5]))
        
        # Full xpath for the "Add an audio track" input
        upload_css = 'input[type=file]'
        upload_xpath = '/html/body/div[28]/div/div[2]/div[1]/div[3]/div/div/div[1]/div/div/div[3]/div/input'
        div_class_str = 'flyout dialog-editor-block-audio yui3-dd-draggable visible'
        xpath_list = self.audio_xpath_list + [
                                              self.get_div('yui3-widget yui3-uploader sqs-uploader'),
                                              self.get_div('sqs-uploader-content'),
                                              'input'
                                              ]
        self.upload_xpath = "//{}".format('/'.join(xpath_list[10:]))
    
    
    
    def browse_to_sermon_archive(self, driver):
        
        # Bring up the browser to the login page
        self.driver_get_url(driver, url_str=self.squarespace_url)
        
        # Fill in the name and password
        self.fill_in_field(driver, field_name='email', field_value='dave.babbitt@gmail.com')
        self.fill_in_field(driver, field_name='password', field_value='Genesis11')
        
        # Click the login button
        self.click_the_login_button(driver)
    
    
    
    def click_the_edit_button(self, driver):
        self.move_to_title(driver)
        print('Clicking the edit button')
        try:
            
            # Wait for button to show up
            button_tag = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, self.button_xpath))
                )
            button_tag.click()
        
        except Exception as e:
            message = str(e).strip()
            print('Waiting for the edit button to be clickable: {}'.format(message))
            while self.obscure_regex.search(message):
                match_obj = self.obscure_regex.search(message)
                if match_obj:
                    self.unobscure_element(driver, match_obj)
                    self.move_to_title(driver)
                    try:
                        
                        # Wait for button to show up
                        button_tag = WebDriverWait(driver, 20).until(
                            EC.element_to_be_clickable((By.XPATH, self.button_xpath))
                            )
                        button_tag.click()
                        message = ""
                        
                    except Exception as e:
                        message = str(e).strip()
                        print('Waiting for the edit button to be clickable (unobscured): {}'.format(message))
    
    
    
    def click_the_login_button(self, driver):
        print('Clicking the login button')
        try:
            
            # Wait for button to show up
            login_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, self.login_css))
                )
            login_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the login button to show up: {}'.format(message))
    
    
    
    def click_web_element(self, driver, xpath, verbose=True):
        if verbose:
            print('Clicking {}'.format(xpath))
        try:
            web_element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, xpath))
                )
            ActionChains(driver).move_to_element(web_element).perform()
            web_element.click()
        except Exception as e:
            message = str(e).strip()
            if verbose:
                print('Waiting for the web element to be visible: {}'.format(message))
            if ('obscures' in message) or ('Other element would receive the click' in message):
                tag_list = obscure_regex.findall(message)
                while len(tag_list) == 2:
                    unobscure_element(driver, tag_list[1])
                    try:
                        web_element = WebDriverWait(driver, 5).until(
                            EC.visibility_of_element_located((By.XPATH, xpath))
                            )
                        ActionChains(driver).move_to_element(web_element).perform()
                        web_element.click()
                        message = ''
                    except Exception as e:
                        message = str(e).strip()
                        if verbose:
                            print('Waiting for the web element to be visible (unobscured): {}'.format(message))
                    tag_list = obscure_regex.findall(message)
    
    
    
    def create_keywords_file(self, test_name='test_one',
                             keyword_list=['Bedsheets', 'Clocks', 'Padlocks'],
                             verbose=True):
        wb = Workbook()
        xlsx_dir = os.path.join('../saves', 'xlsx')
        os.makedirs(name=xlsx_dir, exist_ok=True)
        
        # grab the active worksheet
        ws = wb.active
        
        # Data can be assigned directly to cells
        ws['A1'] = 'Search Term'
        ws['A1'].font = Font(bold=True)
        column_width = 11.14
        ws.column_dimensions['A'].width = column_width
        for i, keyword_str in enumerate(keyword_list):
            ws['A{}'.format(i+2)] = keyword_str
        
        # Save the file
        file_path = os.path.join(xlsx_dir, '{}.xlsx'.format('_'.join([test_name, 'keywords'])))
        if verbose:
            print('Saving to {}'.format(os.path.abspath(file_path)))
        wb.save(file_path)
        
        return file_path
    
    
    
    
    def driver_get_url(self, driver, url_str, verbose=True):
        if verbose:
            print('Getting URL: {}'.format(url_str))
        finished = 0
        fails = 0
        while finished == 0 and fails < 8:
            try:
                driver.get(url_str)
                finished = 1
            except Exception as e:
                message = str(e).strip()
                if verbose:
                    print(message)
                fails = fails + 1
                
                # Wait for 5 seconds
                self.wait_for(5, verbose=verbose)
    
    
    
    
    def fill_in_field(self, driver, field_name, field_value, verbose=True):
        if verbose:
            print('Filling in the {} field with {}'.format(field_name, field_value))
        try:
            input_css = 'input[name="{}"]'.format(field_name)
            
            # Wait for input field to show up
            input_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, input_css))
                )
            input_tag.click()
            ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            input_tag.send_keys(field_value)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for input field to show up: {}'.format(message))
    
    
    
    def get_contains_list(self, div_class_list):
        
        return ["contains(@class, '{}')".format(c) for c in div_class_list]
    
    
    
    def get_div(self, div_class_str):
        
        return 'div[{}]'.format(' and '.join(self.get_contains_list(div_class_str.split(' '))))
    
    
    
    def get_driver(self, browser_name='FireFox'):
        print('Getting the {} driver'.format(browser_name))
        log_dir = '../log'
        os.makedirs(name=log_dir, exist_ok=True)
        if browser_name == 'FireFox':
            executable_name = 'geckodriver'
        elif browser_name == 'Chrome':
            executable_name = 'chromedriver80'
        executable_path = '../../web-scrapers/exe/{}.exe'.format(executable_name)
        service_log_path = os.path.join(log_dir, '{}.log'.format(executable_name))
        if browser_name == 'FireFox':
            fp = webdriver.FirefoxProfile()
            #fp.set_preference(key, value)
            driver = webdriver.Firefox(
                capabilities=None,
                executable_path=executable_path,
                firefox_binary=None,
                firefox_options=None,
                firefox_profile=fp,
                options=None,
                proxy=None,
                service_log_path=service_log_path,
                timeout=30,
                )
        elif browser_name == 'Chrome':
            co = webdriver.ChromeOptions()
            co.add_argument('--no-sandbox')
            #co.set_capability(name, value)
            driver = webdriver.Chrome(
                chrome_options=None,
                executable_path=executable_path,
                keep_alive=True,
                options=co,
                port=0,
                service_log_path=service_log_path,
            )
        
        # Set timeout information
        driver.set_page_load_timeout(20)
        
        return driver
    
    
    
    def get_element_contents(self, driver, xpath, verbose=True, scroll=True):
        if verbose:
            print('Getting text of {}'.format(xpath))
        results_str = ''
        try:
            if scroll:
                driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
            web_element = self.get_web_element(driver, xpath)
            if scroll:
                driver.execute_script('arguments[0].scrollIntoView(true)', web_element)
            results_str = web_element.text.strip()
        except Exception as e:
            message = str(e).strip()
            if verbose:
                print('Waiting for the web element to be located: {}'.format(message))
            if ('obscures' in message) or ('Other element would receive the click' in message):
                tag_list = obscure_regex.findall(message)
                while len(tag_list) == 2:
                    unobscure_element(driver, tag_list[1])
                    try:
                        if scroll:
                            driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
                        web_element = self.get_web_element(driver, xpath)
                        if scroll:
                            driver.execute_script('arguments[0].scrollIntoView(true)', web_element)
                        results_str = web_element.text.strip()
                        message = ''
                    except Exception as e:
                        message = str(e).strip()
                        if verbose:
                            print('Waiting for the web element to be located (unobscured): {}'.format(message))
                    tag_list = obscure_regex.findall(message)
        
        return results_str
    
    
    
    def get_last_sunday(self, today=datetime.now()):
        start = today - timedelta((today.weekday() + 1) % 7)
        
        return start + relativedelta.relativedelta(weekday=relativedelta.SU(-1))
    
    
    
    def get_recent_sermons_dataframe(self, driver, verbose=True):
        NOW_STAMP = datetime.now()
        div_xpath = '/html/body/div[1]/div/div[1]/div/div/div/div/div[2]/div[1]'
        div_tag = self.get_web_element(driver, div_xpath)
        rows_list = []
        columns_list = ['sermon_reference', 'sermon_title', 'publish_date',
                        'speaker_name', 'on_sermonaudio', 'mp3_url', 'mp3_path']
        driver.maximize_window()
        for h2_tag in div_tag.find_elements_by_tag_name('h2'):
            
            # Get the grandparent of the h2 tag
            parent_tag = h2_tag.find_element_by_xpath('..')
            grandparent_tag = parent_tag.find_element_by_xpath('..')
            info_list = grandparent_tag.text.strip().split('\n')
            
            # Get the sermon title
            reference_title_list = info_list[0].split(': ')
            if len(reference_title_list) > 1:
                sermon_reference = reference_title_list[0]
                sermon_title = reference_title_list[1]
            else:
                sermon_reference = np.nan
                sermon_title = reference_title_list[0]
            match_series = (self.recent_sermons_df.sermon_reference == sermon_reference)
            match_series = match_series & (self.recent_sermons_df.sermon_title == sermon_title)
            if self.recent_sermons_df[match_series].shape[0] == 0:
                row_dict = {}
                row_dict['sermon_reference'] = sermon_reference
                row_dict['sermon_title'] = sermon_title
                
                # Get the published date
                publisher_list = info_list[1].split('·')
                date_str = publisher_list[0].strip()
                if len(date_str.split(',')) == 1:
                    
                    # Yesterday 12:00pm
                    if ('Yesterday' in date_str):
                        date_format = 'Yesterday %I:%M%p'
                        publish_date = datetime.strptime(date_str, date_format)
                        last_sunday = self.get_last_sunday(NOW_STAMP)
                        publish_date = publish_date.replace(year=last_sunday.year,
                                                            month=last_sunday.month,
                                                            day=last_sunday.day)
                    
                    # Last Sunday 12:00pm
                    elif ('Last' in date_str):
                        date_format = 'Last %A %I:%M%p'
                        publish_date = datetime.strptime(date_str, date_format)
                        last_sunday = self.get_last_sunday(NOW_STAMP)
                        publish_date = publish_date.replace(year=last_sunday.year,
                                                            month=last_sunday.month,
                                                            day=last_sunday.day)
                    
                    # Dec 22
                    else:
                        date_format = '%b %d'
                        publish_date = datetime.strptime(date_str, date_format)
                        publish_date = publish_date.replace(year=NOW_STAMP.year,
                                                            month=publish_date.month,
                                                            day=publish_date.day)
                        if publish_date > NOW_STAMP:
                            publish_date = publish_date.replace(year=publish_date.year-1)
                            
                row_dict['publish_date'] = publish_date
                
                rows_list.append(row_dict)
            
        self.recent_sermons_df = self.recent_sermons_df.append(pd.DataFrame(rows_list,
                                                                            columns=columns_list),
                                                               ignore_index=True)
        
        # Get the speaker and mp3_url
        div_xpath = '/html/body/div[1]/div/div[1]/div/div/div/div/div[2]/div[1]/div[3]/div/div[1]'
        iframe_xpath = '/html/body/div[1]/div/div[3]/div/div/div[2]/div/div/iframe'
        speaker_xpath = '/html/body/div[4]/main/div/div/article/footer/p/a'
        mp3_xpath = '/html/body/div[4]/main/div/div/article/div/div/div/div/div/div/div/div/div/div/div[3]/div[2]/a'
        match_series = self.recent_sermons_df.speaker_name.isnull()
        match_series = match_series | self.recent_sermons_df.mp3_url.isnull()
        for row_index, row_series in self.recent_sermons_df[match_series].iterrows():
            sermon_reference = row_series['sermon_reference']
            sermon_title = row_series['sermon_title']
            field_value = '{}: {}'.format(sermon_reference, sermon_title)
            driver.switch_to.parent_frame()
            self.search_the_sermon_archive(driver, field_value, verbose=verbose)
            self.wait_for(7, verbose=verbose)
            div_tag = self.get_web_element(driver, div_xpath)
            if div_tag is not None:
                div_tag.click()
                
                # An expectation for checking whether the given frame is available to
                # switch to. If the frame is available it switches the given driver to the
                # specified frame.
                ec_tuple = (By.XPATH, iframe_xpath)
                ec_type = EC.frame_to_be_available_and_switch_to_it(ec_tuple)
                WebDriverWait(driver, 10).until(ec_type)
                
                speaker_tag = self.get_web_element(driver, speaker_xpath)
                if speaker_tag is not None:
                    self.recent_sermons_df.loc[row_index, 'speaker_name'] = speaker_tag.text.strip()
                
                mp3_tag = self.get_web_element(driver, mp3_xpath)
                if mp3_tag is not None:
                    self.recent_sermons_df.loc[row_index, 'mp3_url'] = mp3_tag.get_attribute('href').strip()
        
        # Fix Andrew's name
        speaker_name = 'Andrew Davis'
        full_name = 'Andrew Franklin Davis'
        match_series = (self.recent_sermons_df.speaker_name == speaker_name)
        self.recent_sermons_df.loc[match_series, 'speaker_name'] = full_name
        
        # Get whether it's already on sermonaudio.com
        match_series = ~self.recent_sermons_df['speaker_name'].isnull()
        match_series = match_series & self.recent_sermons_df['on_sermonaudio'].isnull()
        for row_index, row_series in self.recent_sermons_df[match_series].iterrows():
            speaker_name = row_series['speaker_name']
            sermon_title = row_series['sermon_title']
            title_str = '{}'.format(sermon_title)
            on_sermonaudio = self.on_sermonaudio(driver, speaker_name, title_str, verbose=verbose)
            self.recent_sermons_df.loc[row_index, 'on_sermonaudio'] = on_sermonaudio
        
        # Download the mp3
        mp3_dir = os.path.join(self.s.data_folder, 'mp3')
        match_series = ~self.recent_sermons_df['mp3_url'].isnull()
        match_series = match_series & self.recent_sermons_df['mp3_path'].isnull()
        for row_index, row_series in self.recent_sermons_df[match_series].iterrows():
            mp3_url = row_series['mp3_url']
            file_name = mp3_url.split('?')[0].split('/')[-1]
            file_path = os.path.join(mp3_dir, file_name)
            if Path(file_path).is_file():
                self.recent_sermons_df.loc[row_index, 'mp3_path'] = Path(file_path).absolute().resolve()
            else:
                try:
                    message_tuple = urlretrieve(url=mp3_url, filename=file_path)
                    self.recent_sermons_df.loc[row_index, 'mp3_path'] = Path(message_tuple[0]).absolute().resolve()
                except Exception as e:
                    if verbose:
                        print('Error on {} ({}): {}'.format(mp3_url, file_path, e))
        
        return self.recent_sermons_df
    
    
    
    def get_web_element(self, driver, xpath):
        try:
            web_element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, xpath))
                )
            
            return web_element
            
        except Exception as e:
            message = str(e).strip()
            print('Waiting for the web_element to show up: {}'.format(message))
    
    
    
    def key_in_search(self, driver, input_xpath, field_value, verbose=True):
        if verbose:
            print('Searching for: {}'.format(field_value))
        input_tag = self.get_web_element(driver, input_xpath)
        input_tag.click()
        ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
        input_tag.send_keys(field_value)
        ActionChains(driver).key_down(Keys.ENTER).perform()
    
    
    
    def log_into_squarespace(self, driver, verbose=True):
        self.driver_get_url(driver, self.squarespace_url, verbose=verbose)
        self.fill_in_field(driver, field_name='email',
                           field_value='dave.babbitt@gmail.com',
                           verbose=verbose)
        self.fill_in_field(driver, field_name='password',
                           field_value='202001squarespace@WSXcde3$RFVbgt5',
                           verbose=verbose)
        button_xpath = '/html/body/div[1]/div/div[2]/div/div[1]/div/div[1]/div[1]/div/button'
        self.click_web_element(driver, xpath=button_xpath, verbose=verbose)
        
        # Wait for 10 seconds
        self.wait_for(10, verbose=verbose)
    
    
    
    def move_to(self, driver, xpath, verbose=True):
        if verbose:
            print('Moving to {}'.format(xpath))
        try:
            web_element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, xpath))
                )
            ActionChains(driver).move_to_element(web_element).perform()
        except Exception as e:
            message = str(e).strip()
            if verbose:
                print('Waiting for the web element to be visible: {}'.format(message))
            if 'obscures' in message:
                match_obj = obscure_regex.search(message)
                while match_obj:
                    unobscure_element(driver, match_obj)
                    try:
                        web_element = WebDriverWait(driver, 5).until(
                            EC.visibility_of_element_located((By.XPATH, xpath))
                            )
                        ActionChains(driver).move_to_element(web_element).perform()
                        message = ''
                    except Exception as e:
                        message = str(e).strip()
                        if verbose:
                            print('Waiting for the web element to be visible (unobscured): {}'.format(message))
                    match_obj = obscure_regex.search(message)
    
    
    
    def move_to_title(self, driver):
        print('Moving to the title')
        status_tag = self.get_web_element(driver, self.title_xpath)
        ActionChains(driver).move_to_element(status_tag).click().perform()
    
    
    
    def search_the_sermon_archive(self, driver, field_value, verbose=True):
        if verbose:
            print('Filling in the search field with "{}"'.format(field_value))
        try:
            
            # Wait for input field to show up
            self.key_in_search(driver, self.input_xpath, field_value, verbose=verbose)
            
        except Exception as e:
            message = str(e).strip()
            if verbose:
                print('Waiting for search field to show up (first attempt): {}'.format(message))
            if 'scrolled' in message:
                driver.refresh()
                
                # Wait for 10 seconds
                self.wait_for(10, verbose=verbose)
                
                try:
                    
                    # Wait for input field to show up
                    self.key_in_search(driver, self.input_xpath, field_value, verbose=verbose)
                    
                except Exception as e:
                    message = str(e).strip()
                    if verbose:
                        print('self.key_in_search (second attempt): {}'.format(message))
            else:
                while self.obscure_regex.search(message):
                    match_obj = self.obscure_regex.search(message)
                    if match_obj:
                        self.unobscure_element(driver, match_obj)
                        try:
                            
                            # Wait for input field to show up
                            self.key_in_search(driver, self.input_xpath, field_value, verbose=verbose)
                            message = ''
                            
                        except Exception as e:
                            message = str(e).strip()
                            if verbose:
                                print('Waiting for search field to show up (unobscured): {}'.format(message))
    
    
    
    def similar(self, a, b):
        
        return SequenceMatcher(None, str(a), str(b)).ratio()
    
    
    
    def unobscure_element(self, driver, match_tuple):
        tag_name = match_tuple[0]
        class_str = '.'.join([cn for cn in match_tuple[1].split(' ') if cn != 'hidden'])
        overlay_css = '{}.{}'.format(tag_name, class_str)
        try:
            
            # Wait for overlay div to show up
            overlay_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, overlay_css))
                )
            print('Hiding <{}'.format(overlay_tag.get_attribute('outerHTML').split('<')[1]))
            driver.execute_script("arguments[0].setAttribute('style','display:none;');", overlay_tag)
        
        except Exception as e:
            message = str(e).strip()
            print('Waiting for {} {} to show up before hiding it: {}'.format(class_str, tag_name, message))
            driver.refresh()
    
    
    
    def wait_for(self, wait_count, verbose=True):
        if verbose:
            print('Waiting for {} seconds'.format(wait_count))
        time.sleep(wait_count)
    
    
    
    def click_the_audio_apply_button(self, driver):
        apply_css = 'input[tabindex="117"]'
        apply_css = 'div.controls-block > div.button-holder > div.button-block > input'
        apply_xpath = '/html/body/div[28]/div/div[3]/div[3]/div[2]/input'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui flyout light buttons-right dialog-editor-block-audio yui3-dd-draggable visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div", "div[3]", "div[3]", "div[2]", "input"]
        apply_xpath = "/{}".format('/'.join(xpath_list))
        apply_tag = self.get_web_element(driver, apply_xpath)
        apply_tag.click()
    
    
    
    def add_an_episode_note(self, driver, note_str):
        self.click_the_podcasting_tab(driver)
        note_css = 'div.field-rhs > div.field-rte-wrapper > div > div.rte.ProseMirror > p'
        note_xpath = '/html/body/div[27]/div/div[2]/div[2]/div[5]/div/div[2]/div[1]/div/div[2]'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui flyout light buttons-right dialog-editor-block-audio yui3-dd-draggable visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div", "div[2]", "div[2]", "div[5]", "div", "div[2]", "div[1]", "div", "div[2]"]
        note_xpath = "/{}".format('/'.join(xpath_list))
        print('Clicking the "Episode Note" text area and pasting in "{}"'.format(note_str))
        note_tag = self.get_web_element(driver, note_xpath)
        ActionChains(driver).move_to_element(note_tag).perform()
        note_tag.click()
        ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
        ActionChains(driver).key_down(Keys.CONTROL).send_keys('x').key_up(Keys.CONTROL).perform()
        note_tag.send_keys(note_str)
    
    
    
    def click_the_edit_overlay(self, driver):
        div_class_str = 'yui3-widget sqs-widget sqs-data-widget sqs-dialog-field sqs-button-vanilla normal-button data-state-loaded'
        edit_css = 'div[class="{}"][data-test="button"][title="Edit"]'.format(div_class_str)
        edit_xpath = '/html/body/div[28]/div[1]/div[2]/div[1]/div[2]/div[1]/div/div[1]/div/div[1]/div/div[2]/div[3]/div/div[2]'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui standard light buttons-right dialog-text-post-editor visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div[1]", "div[2]", "div[1]", "div[2]", "div[1]", "div", "div[1]", "div",
                      "div[1]", "div", "div[2]", "div[3]", "div", "div[2]"]
        edit_xpath = "/{}".format('/'.join(xpath_list))
        try:
            print('Clicking the "EDIT" overlay')
            edit_tag = WebDriverWait(driver, 10).until(
                #EC.presence_of_element_located((By.XPATH, edit_xpath))
                EC.presence_of_element_located((By.CSS_SELECTOR, edit_css))
                )
            ActionChains(driver).move_to_element(edit_tag).click().perform()
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the "EDIT" overlay to show up: {}'.format(message))
    
    
    def click_the_podcasting_tab(self, driver):
        tab_css = 'div[data-test="tab-1"]'
        tab_xpath = '/html/body/div[27]/div/div[1]/div[2]/div/div/div/div[2]'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui flyout light buttons-right dialog-editor-block-audio yui3-dd-draggable visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div", "div[1]", "div[2]", "div", "div", "div", "div[2]"]
        tab_xpath = "/{}".format('/'.join(xpath_list))
        print('Clicking the "Podcasting" tab')
        tab_tag = self.get_web_element(driver, tab_xpath)
        tab_tag.click()
    
    
    
    def click_the_save_post_button(self, driver):
        save_css = '.saveAndClose'
        save_css_list = ['div.controls-block', 'div.button-holder', 'div:nth-child(4)',
                         'input[class="saveAndClose"][tabindex="105"][type="button"][data-test="dialog-saveAndClose"][value="Save"]']
        save_css = ' > '.join(save_css_list)
        save_xpath = '/html/body/div[58]/div[1]/div[3]/div[3]/div[4]/input'
        save_xpath = '/html/body/div[85]/div[1]/div[3]/div[3]/div[4]/input'
        save_xpath = '/html/body/div[15]/div[1]/div[3]/div[3]/div[4]/input'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui standard light buttons-right dialog-text-post-editor visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div[1]", "div[3]", "div[3]", "div[4]", "input"]
        save_xpath = "/{}".format('/'.join(xpath_list))
        print('Clicking the "SAVE" button')
        save_tag = self.get_web_element(driver, save_xpath)
        save_tag.click()
    
    
    
    def click_the_upload_file_subtab(self, driver):
        tab_css = 'div[class="option-title"]'
        tab_xpath = '/html/body/div[28]/div/div[2]/div[1]/div[2]/div/div[1]/div/div'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui flyout light buttons-right dialog-editor-block-audio yui3-dd-draggable visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div", "div[2]", "div[1]", "div[2]", "div", "div[1]", "div", "div"]
        tab_xpath = "/{}".format('/'.join(xpath_list))
        print('Clicking the "Upload File" subtab')
        tab_tag = self.get_web_element(driver, tab_xpath)
        tab_tag.click()
    
    
    
    def click_external_file_subtab(self, driver):
        print('Clicking the external file tab')
        button_css = 'div[data-value="external"]'
        try:
            
            # Wait for button to show up
            button_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, button_css))
            )
            button_tag.click()
            
        except Exception as e:
            raise Exception('Waiting for the external file tab to show up: {}'.format(e))
    
    
    
    def click_the_close_signin_window_link(self, driver):
        print('Clicking the "Click here" link')
        #clickhere_css = 'body > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(1) > p:nth-child(2) > font:nth-child(1)
        #> a:nth-child(2)'
        #clickhere_xpath = '/html/body/table/tbody/tr[2]/td/p/font/a/b'
        clickhere_css = 'a[href="javascript:void(close())"]'
        try:
            
            # Wait for signin button to show up
            clickhere_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, clickhere_css))
            )
            clickhere_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the "Click here" link to show up: {}'.format(message))
    
    
    
    def click_the_member_signin_button(self, driver):
        print('Clicking the "Member Sign In" button')
        #signin_css = 'body > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(1) > form:nth-child(2) > table:nth-child(1)
        #> tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1) > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(4) > td:nth-child(2)
        #> input:nth-child(7)'
        #signin_xpath = '/html/body/table[1]/tbody/tr/td/table/tbody/tr/td[4]/font/a'
        signin_css = 'input[value="Member Sign In"]'
        try:
            
            # Wait for signin button to show up
            signin_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, signin_css))
                )
            signin_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for signin button to show up: {}'.format(message))
    
    
    
    def click_the_signin_button(self, driver):
        #signin_css = '.navleft3e'
        #signin_xpath = '/html/body/table[1]/tbody/tr/td/table/tbody/tr/td[4]/font/a'
        print('Clicking the "Sign in or signup" link')
        signin_css = 'a[href="{}"]'.format(self.signin_url)
        try:
            main_window_handle = None
            while not main_window_handle:
                main_window_handle = driver.current_window_handle
            
            # Wait for signin link to show up
            signin_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, signin_css))
                )
            signin_tag.click()
            signin_window_handle = None
            while not signin_window_handle:
                for handle in driver.window_handles:
                    if handle != main_window_handle:
                        signin_window_handle = handle
                        break
            driver.switch_to.window(signin_window_handle)
            username = 'redeemer-ma-pca'
            self.fill_in_field(driver, 'Source', username)
            password = '191SudburyRoad'
            self.fill_in_field(driver, 'Password', password)
            self.click_the_member_signin_button(driver)
            self.click_the_close_signin_window_link(driver)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for signin link to show up: {}'.format(message))
        
        return main_window_handle
    
    
    
    def get_the_file_preview(self, driver):
        print('Getting the file preview')
        preview_tag = self.get_web_element(driver, self.preview_xpath)
        preview_str = preview_tag.text.strip().split('\n')[0]
        
        return preview_str
    
    
    
    def on_sermonaudio(self, driver, speaker_str, title_str, verbose=True):
        if verbose:
            print('Checking if this sermon is already uploaded to sermonaudio.com')
        keyword = '_'.join(speaker_str.split(' '))
        keyword_within = '+'.join(title_str.split(' '))
        search_url = self.search_formatted_str.format(keyword, keyword_within)
        self.driver_get_url(driver, search_url, verbose=verbose)
        search_css = 'font.ar2 > font.ve2'
        try:
            
            # Wait for count to show up
            font_tag = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, search_css))
                )
            sermon_count = int(font_tag.find_elements_by_tag_name('b')[0].text)
            
        except Exception as e:
            message = str(e).strip()
            if verbose:
                print('Waiting for the count to show up: {}'.format(message))
            sermon_count = 0
        
        return sermon_count > 0



    def click_the_submit_button(self, driver, verbose=True):
        if verbose:
            print('Clicking the submit sermon button')
        button_css = 'input[name="submitsermon"]'
        button_css = '#topbarBackgroundImage > table > tbody > tr > td > form > table:nth-child(3) > tbody > tr:nth-child(13) > td:nth-child(2) > p > input[type=SUBMIT]:nth-child(3)'
        button_xpath = '/html/body/table[2]/tbody/tr/td/table/tbody/tr[3]/td[2]/div/div/div/div/table/tbody/tr/td/table/tbody/tr/td/form/table[2]/tbody/tr[13]/td[2]/p/input[2]'
        try:
            
            # Wait for button to show up
            self.move_to(driver, button_xpath, verbose=verbose)
            button_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, button_css))
                )
            button_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the button to show up: {}'.format(message))



    def click_the_copyright_checkbox(self, driver, verbose=True):
        if verbose:
            print('Clicking the copyright checkbox')
        checkbox_css = 'input[name="copyright"]'
        checkbox_css = '#copyright'
        checkbox_xpath = '/html/body/table[2]/tbody/tr/td/table/tbody/tr[3]/td[2]/div/div/div/div/table/tbody/tr/td/table/tbody/tr/td/form/table[2]/tbody/tr[13]/td[2]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/input'
        try:
            
            # Wait for checkbox to show up
            self.move_to(driver, checkbox_xpath, verbose=verbose)
            checkbox_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, checkbox_css))
                )
            checkbox_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the checkbox to show up: {}'.format(message))



    def fill_in_textarea(self, driver, field_name, field_value):
        print('Filling in the {} field with {}'.format(field_name, field_value))
        try:
            input_css = 'textarea[name="{}"]'.format(field_name)
            
            # Wait for textarea field to show up
            input_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, input_css))
                )
            input_tag.click()
            ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            input_tag.send_keys(field_value)
        
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for textarea field to show up: {}'.format(message))



    def fill_in_sermon_info(self, driver, description_str, title_str, speaker_str, reference_str, datetime_obj):
        print('Filling in the sermon info')
        self.driver_get_url(driver, self.submitsermon_url)
        self.fill_in_field(driver, 'SourceID', 'redeemer-ma-pca')
        self.fill_in_field(driver, 'SourcePassword', 'Genesis11')
        main_window_handle = self.click_the_speaker_selector_button(driver)
        self.fill_in_field(driver, 'keyword', speaker_str)
        self.click_the_search_button(driver)
        self.click_the_first_link(driver, speaker_str)
        
        driver.switch_to.window(main_window_handle)
        self.fill_in_field(driver, 'Title', title_str)
        self.fill_in_field(driver, 'Date', datetime_obj.strftime('%m/%d/%Y'))
        self.fill_in_field(driver, 'BibleText', reference_str)
        if description_str != '':
            self.fill_in_textarea(driver, 'moreinfotext', description_str)
        driver.switch_to.window(main_window_handle)
        self.click_the_copyright_checkbox(driver)
        self.click_the_submit_button(driver)
        
        return main_window_handle



    def get_the_sermon_id(self, driver, main_window_handle):
        print('Getting the sermon ID')
        id_css = 'td:nth-child(2) > font[class="ve5b"]'
        try:
            driver.switch_to.window(main_window_handle)
            
            # Wait for ID to show up
            font_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, id_css))
                )
            sermon_id = font_tag.find_elements_by_tag_name('b')[0].text
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the ID to show up: {}'.format(message))
        
        return sermon_id



    def click_the_upload_media_button(self, driver, main_window_handle):
        print('Clicking the upload media button')
        upload_css = 'font.ar2 > a.navleftblack5b'
        try:
            driver.switch_to.window(main_window_handle)
            
            # Wait for button to show up
            upload_tag = WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, upload_css))
                )
            upload_tag.click()
            upload_window_handle = None
            while not upload_window_handle:
                for handle in driver.window_handles:
                    if handle != main_window_handle:
                        upload_window_handle = handle
                        break
            driver.switch_to.window(upload_window_handle)
            for frame_tag in driver.find_elements_by_tag_name('frame'):
                if frame_tag.get_attribute('name') == 'AudioAdd1':
                    driver.switch_to.frame(frame_tag)
                    break
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the ID to show up: {}'.format(message))
        
        return upload_window_handle



    def fill_in_the_password_field(self, driver, upload_window_handle):
        print('Filling in the password field')
        password_css = '#password'
        password_xpath = '//*[@id="password"]'
        try:
            
            # Wait for input to show up
            driver.switch_to.window(upload_window_handle)
            driver.maximize_window()
            for frame_tag in driver.find_elements_by_tag_name('frame'):
                if frame_tag.get_attribute('name') == 'AudioAdd1':
                    driver.switch_to.frame(frame_tag)
                    break
            password_tag = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, password_xpath))
                )
            password_tag.click()
            ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            password_tag.send_keys('Genesis11')
            
        except Exception as e:
            message = str(e).strip()
            print('Waiting for the password field to show up (first time): {}'.format(message))
            try:
                
                # Wait for input to show up
                driver.switch_to.window(upload_window_handle)
                driver.maximize_window()
                for frame_tag in driver.find_elements_by_tag_name('frame'):
                    if frame_tag.get_attribute('name') == 'AudioAdd1':
                        driver.switch_to.frame(frame_tag)
                        break
                password_tag = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, password_css))
                    )
                password_tag.click()
                ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
                password_tag.send_keys('Genesis11')
                
            except Exception as e:
                message = str(e).strip()
                raise Exception('Waiting for the password field to show up (second time): {}'.format(message))



    def click_the_browse_button(self, driver, file_path):
        print('Clicking the browse button')
        browse_css = 'input[name="datafile"]'
        try:
            
            # Wait for button to show up
            browse_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, browse_css))
            )
            browse_tag.send_keys(file_path)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the browse window to show up: {}'.format(message))
    
    
    
    def click_another_upload_media_button(self, driver):
        print('Clicking the secondary upload media button')
        upload_css = 'input[name="action_upload"]'
        try:
            
            # Wait for button to show up
            upload_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, upload_css))
                )
            upload_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the button to show up: {}'.format(message))



    def click_the_close_upload_window_link(self, driver, main_window_handle):
        print('Uploading the MP3 and waiting for processing...')
        upload_window_handle = None
        while not upload_window_handle:
            for handle in driver.window_handles:
                if handle != main_window_handle:
                    upload_window_handle = handle
                    break
        driver.switch_to.window(upload_window_handle)
        for frame_tag in driver.find_elements_by_tag_name('frame'):
            if frame_tag.get_attribute('name') == 'AudioAdd1':
                driver.switch_to.frame(frame_tag)
                break
        close_css = 'body > table > tbody > tr:nth-child(2) > td > font > p:nth-child(6) > a'
        try:
            close_tag = WebDriverWait(driver, 140).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, close_css))
                )
            close_tag.click()
            driver.switch_to.window(main_window_handle)
        except Exception as e:
            message = str(e).strip()
            print('Waiting for the close window link to show up (first attempt): {}'.format(message))
            close_xpath = '/html/body/table/tbody/tr[2]/td/font/p[3]/a'
            try:
                close_tag = WebDriverWait(driver, 140).until(
                    EC.presence_of_element_located((By.XPATH, close_xpath))
                    )
                close_tag.click()
                driver.switch_to.window(main_window_handle)
            except Exception as e:
                message = str(e).strip()
                print('Waiting for the close window link to show up (second attempt): {}'.format(message))
                try:
                    driver.execute_script('top.close();')
                    driver.switch_to.window(main_window_handle)
                except Exception as e:
                    message = str(e).strip()
                    raise Exception('Executing top.close() javascript (final attempt): {}'.format(message))



    def check_for_audio(self, driver, row_index, main_window_handle):
        print('Checking for the audio')
        audio_css = 'font.ve2 > div > a > b'
        try:
            
            # Wait for the audio link to show up
            driver.switch_to.window(main_window_handle)
            audio_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, audio_css))
                )
            if 'Audio' in audio_tag.text:
                self.recent_sermons_df.loc[row_index, 'on_sermonaudio'] = True
            else:
                self.recent_sermons_df.loc[row_index, 'on_sermonaudio'] = False
            self.s.store_objects(recent_sermons_df=self.recent_sermons_df)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the audio link to show up: {}'.format(message))



    def cancel_upload_window(self, driver, upload_window_handle, upload_url):
        print('Clicking the cancel upload button')
        cancel_css = 'input[name="action_cancel"]'
        try:
            
            # Wait for button to show up
            driver.switch_to.window(upload_window_handle)
            cancel_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, cancel_css))
                )
            cancel_tag.click()
            if driver.current_url == upload_url:
                driver.close()
            
        except Exception as e:
            message = str(e).strip()
            print('Waiting for the cancel upload button to show up: {}'.format(message))



    def click_the_edit_media_button(self, driver, edit_url, main_window_handle):
        print('Clicking the edit media button')
        edit_css = 'a[href="{}"'.format(edit_url)
        try:
            
            # Wait for button to show up
            edit_tag = WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, edit_css))
                )
            edit_tag.click()
            edit_window_handle = None
            while not edit_window_handle:
                for handle in driver.window_handles:
                    if handle != main_window_handle:
                        edit_window_handle = handle
                        break
            driver.switch_to.window(edit_window_handle)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the ID to show up: {}'.format(message))
        
        return edit_window_handle



    def click_the_delete_entry_button(self, driver, edit_url, edit_window_handle, main_window_handle):
        print('Clicking the delete media button')
        edit_url = self.formatted_str.format(edit_url)
        delete_css = 'input[name="deletebutton"]'
        try:
            
            # Wait for button to show up
            delete_tag = WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, delete_css))
            )
            delete_tag.click()
            alert = driver.switch_to.alert
            alert.accept()
            driver.switch_to.window(edit_window_handle)
            if driver.current_url == 'https://www.sermonaudio.com/winedit_sermon.asp':
                driver.close()
                driver.switch_to.window(main_window_handle)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the ID to show up: {}'.format(message))
        
        return main_window_handle



    def upload_the_mp3(self, driver, main_window_handle, sermon_id, file_path, row_index):
        print('Uploading the MP3')
        sermon_url = 'http://www.sermonaudio.com/sermoninfo.asp?SID={}'.format(sermon_id)
        self.driver_get_url(driver, sermon_url)
        upload_window_handle = self.click_the_upload_media_button(driver, main_window_handle)
        self.fill_in_the_password_field(driver, upload_window_handle)
        try:
            self.click_the_browse_button(driver, file_path)
            self.click_another_upload_media_button(driver)
            self.click_the_close_upload_window_link(driver, main_window_handle)
            time.sleep(6*60)
            driver.refresh()
            self.check_for_audio(driver, row_index, main_window_handle)
            #os.remove(file_path)
        except Exception as e:
            message = str(e).strip()
            print('Attempting to upload the MP3: {}'.format(message))
            upload_url = 'https://web4.sermonaudio.com/winedit_audioadd-aspupload.asp?sermonID={}'.format(sermon_id)
            self.cancel_upload_window(driver, upload_window_handle, upload_url)
            driver.switch_to.window(main_window_handle)
            edit_url = 'winedit_sermon.asp?SermonID={}'.format(sermon_id)
            edit_window_handle = self.click_the_edit_media_button(driver, edit_url, main_window_handle)
            driver.switch_to.window(edit_window_handle)
            self.fill_in_field(driver, 'password', 'Genesis11')
            self.click_the_delete_entry_button(driver, edit_url, edit_window_handle, main_window_handle)
            driver.switch_to.window(main_window_handle)



    def click_the_speaker_selector_button(self, driver):
        print('Clicking the speaker selector button')
        button_css = 'a[href="winedit_speaker.asp"]'
        try:
            main_window_handle = None
            while not main_window_handle:
                main_window_handle = driver.current_window_handle
            
            # Wait for button to show up
            button_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, button_css))
                )
            button_tag.click()
            signin_window_handle = None
            while not signin_window_handle:
                for handle in driver.window_handles:
                    if handle != main_window_handle:
                        signin_window_handle = handle
                        break
            driver.switch_to.window(signin_window_handle)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for speaker selector button to show up: {}'.format(message))
        
        return main_window_handle



    def click_the_add_new_speaker_button(self, driver):
        print('Clicking the add new speaker button')
        button_css = 'input[value="Add New Speaker"]'
        try:
            
            # Wait for button to show up
            button_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, button_css))
                )
            button_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the seearch button to show up: {}'.format(message))



    def click_the_new_speaker_checkbox(self, driver):
        print('Clicking the new speaker checkbox')
        checkbox_css = 'input[name="speakercheck"]'
        try:
            
            # Wait for checkbox to show up
            checkbox_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, checkbox_css))
                )
            checkbox_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the new speaker checkbox to show up: {}'.format(message))



    def click_the_first_link(self, driver, search_str):
        print('Clicking the first link')
        link_css = 'a.addtocartlink:nth-child(1)'
        try:
            
            # Wait for link to show up
            link_css = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, link_css))
                )
            link_css.click()
            
        except Exception as e:
            message = str(e).strip()
            print('Adding a new speaker ({}): {}'.format(search_str, message))
            self.click_the_new_speaker_checkbox(driver)
            self.fill_in_field(driver, 'AddSpeaker', search_str)
            self.click_the_add_new_speaker_button(driver)



    def click_the_search_button(self, driver):
        print('Clicking the search button')
        button_css = 'input[name="search"]'
        try:
            
            # Wait for button to show up
            button_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, button_css))
            )
            button_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the seearch button to show up: {}'.format(message))



    def get_mp3_filepath(self, mp3_url):
        parts_list = mp3_url.split('/')
        file_name = urllib.parse.quote(parts_list[-1])
        url_prefix = parts_list[:-1]
        mp3_url = '/'.join(url_prefix+[file_name])
        print('Getting MP3 file path for {}'.format(mp3_url))
        mp3_dir = '../data/mp3'
        os.makedirs(name=mp3_dir, exist_ok=True)
        file_path = os.path.abspath(os.path.join(mp3_dir, file_name))
        if not Path(file_path).is_file():
            path_tuple = urlretrieve(url=mp3_url, filename=file_path)
            file_path = path_tuple[0]
        
        return file_path



    def get_speaker_str(self, id_num, library_items_div):
        print('Getting the speaker')
        speaker_selector = '#audioInfo{} > div > div:nth-child(2) > div:nth-child(1)'.format(id_num)
        speaker_str = library_items_div.select(speaker_selector)[0].text.strip()
        speaker_regex = re.compile(r'(Mr\. |Pasor |Pastor |Rev\. |pastor )?([^\r\n]+)')
        speaker_str = speaker_regex.sub(r'\g<2>', speaker_str)
        if speaker_str == 'Kerr':
            speaker_str = 'Matthew Kerr'
        if speaker_str == 'Andrew Davis':
            speaker_str = 'Andrew Franklin Davis'
        
        return speaker_str
