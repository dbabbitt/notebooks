
#!/usr/bin/env python
# Utility Functions to Initialize Simulator Cells.
# Dave Babbitt <dave.babbitt@gmail.com>
# Author: Dave Babbitt, Data Scientist
# coding: utf-8
"""
ScrapingUtilities: A set of utility functions common to web scraping
"""
from datetime import datetime
from datetime import timedelta
from dateutil import relativedelta
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from pathlib import Path
from pydub import AudioSegment
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from shutil import copyfile
from urllib.request import urlretrieve
import io
import math
import numpy as np
import os
import pandas as pd
import random
import re
import storage as s
import time
import urllib
import wikipedia
import warnings
warnings.filterwarnings("ignore")

bs = wikipedia.BeautifulSoup

class ScrapingUtilities(object):
    """This class implements the core of the utility functions
    needed to scrape web content.
    
    Examples
    --------
    
    >>> import scraping_utils
    >>> u = scraping_utils.ScrapingUtilities()
    """
    
    def __init__(self):
        self.s = s.Storage()
        
        # Obscuration error pattern
        self.obscure_regex = re.compile('<([^ ]+)[^>]*class="([^"]+)"[^>]*>')
        
        self.url_regex = re.compile(r'\b(https?|file)://[-A-Z0-9+&@#/%?=~_|$!:,.;]*[A-Z0-9+&@#/%=~_|$]', re.IGNORECASE)
        self.filepath_regex = re.compile(r'\b[c-d]:\\(?:[^\\/:*?"<>|\x00-\x1F]{0,254}[^.\\/:*?"<>|\x00-\x1F]\\)*(?:[^\\/:*?"<>|\x00-\x1F]{0,254}[^.\\/:*?"<>|\x00-\x1F])', re.IGNORECASE)
    
    
    
    def get_page_tables(self, url_or_filepath_or_html, verbose=True):
        if self.url_regex.fullmatch(url_or_filepath_or_html) or self.filepath_regex.fullmatch(url_or_filepath_or_html):
            tables_df_list = pd.read_html(url_or_filepath_or_html)
        else:
            f = io.StringIO(url_or_filepath_or_html)
            tables_df_list = pd.read_html(f)
        if verbose:
            print(sorted([(i, df.shape) for (i, df) in enumerate(tables_df_list)],
                         key=lambda x: x[1][0], reverse=True))
        
        return tables_df_list


    def get_page_soup(self, url_or_filepath_or_html):
        if self.url_regex.fullmatch(url_or_filepath_or_html):
            with urllib.request.urlopen(url_or_filepath_or_html) as response:
                page_html = response.read()
        elif self.filepath_regex.fullmatch(url_or_filepath_or_html):
            with open(url_or_filepath_or_html, 'r', encoding='utf-8') as f:
                page_html = f.read()
        else:
            page_html = url_or_filepath_or_html
        page_soup = bs(page_html, 'html.parser')
        
        return page_soup
    
    
    
    def click_web_element(self, driver, xpath, verbose=True):
        if verbose:
            print('Clicking {}'.format(xpath))
        try:
            web_element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, xpath))
                )
            ActionChains(driver).move_to_element(web_element).perform()
            web_element.click()
        except Exception as e:
            message = str(e).strip()
            if verbose:
                print('Waiting for the web element to be visible: {}'.format(message))
            if ('obscures' in message) or ('Other element would receive the click' in message):
                tag_list = obscure_regex.findall(message)
                while len(tag_list) == 2:
                    unobscure_element(driver, tag_list[1])
                    try:
                        web_element = WebDriverWait(driver, 5).until(
                            EC.visibility_of_element_located((By.XPATH, xpath))
                            )
                        ActionChains(driver).move_to_element(web_element).perform()
                        web_element.click()
                        message = ''
                    except Exception as e:
                        message = str(e).strip()
                        if verbose:
                            print('Waiting for the web element to be visible (unobscured): {}'.format(message))
                    tag_list = obscure_regex.findall(message)
    
    
    
    def create_keywords_file(self, test_name='test_one',
                             keyword_list=['Bedsheets', 'Clocks', 'Padlocks'],
                             verbose=True):
        wb = Workbook()
        xlsx_dir = os.path.join('../saves', 'xlsx')
        os.makedirs(name=xlsx_dir, exist_ok=True)
        
        # grab the active worksheet
        ws = wb.active
        
        # Data can be assigned directly to cells
        ws['A1'] = 'Search Term'
        ws['A1'].font = Font(bold=True)
        column_width = 11.14
        ws.column_dimensions['A'].width = column_width
        for i, keyword_str in enumerate(keyword_list):
            ws['A{}'.format(i+2)] = keyword_str
        
        # Save the file
        file_path = os.path.join(xlsx_dir, '{}.xlsx'.format('_'.join([test_name, 'keywords'])))
        if verbose:
            print('Saving to {}'.format(os.path.abspath(file_path)))
        wb.save(file_path)
        
        return file_path
    
    
    
    
    def driver_get_url(self, driver, url_str, verbose=True):
        if verbose:
            print('Getting URL: {}'.format(url_str))
        finished = 0
        fails = 0
        while finished == 0 and fails < 8:
            try:
                driver.get(url_str)
                finished = 1
            except Exception as e:
                message = str(e).strip()
                if verbose:
                    print(message)
                fails = fails + 1
                
                # Wait for 5 seconds
                self.wait_for(5, verbose=verbose)
    
    
    
    
    def fill_in_field(self, driver, field_name, field_value, verbose=True):
        if verbose:
            print('Filling in the {} field with {}'.format(field_name, field_value))
        try:
            input_css = 'input[name="{}"]'.format(field_name)
            
            # Wait for input field to show up
            input_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, input_css))
                )
            input_tag.click()
            ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            input_tag.send_keys(field_value)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for input field to show up: {}'.format(message))
    
    
    
    def get_contains_list(self, div_class_list):
        
        return ["contains(@class, '{}')".format(c) for c in div_class_list]
    
    
    
    def get_div(self, div_class_str):
        
        return 'div[{}]'.format(' and '.join(self.get_contains_list(div_class_str.split(' '))))
    
    
    
    def get_driver(self, browser_name='FireFox'):
        print('Getting the {} driver'.format(browser_name))
        log_dir = '../log'
        os.makedirs(name=log_dir, exist_ok=True)
        if browser_name == 'FireFox':
            executable_name = 'geckodriver'
        elif browser_name == 'Chrome':
            executable_name = 'chromedriver80'
        executable_path = '../../web-scrapers/exe/{}.exe'.format(executable_name)
        service_log_path = os.path.join(log_dir, '{}.log'.format(executable_name))
        if browser_name == 'FireFox':
            fp = webdriver.FirefoxProfile()
            #fp.set_preference(key, value)
            driver = webdriver.Firefox(
                capabilities=None,
                executable_path=executable_path,
                firefox_binary=None,
                firefox_options=None,
                firefox_profile=fp,
                options=None,
                proxy=None,
                service_log_path=service_log_path,
                timeout=30,
                )
        elif browser_name == 'Chrome':
            co = webdriver.ChromeOptions()
            co.add_argument('--no-sandbox')
            #co.set_capability(name, value)
            driver = webdriver.Chrome(
                chrome_options=None,
                executable_path=executable_path,
                keep_alive=True,
                options=co,
                port=0,
                service_log_path=service_log_path,
            )
        
        # Set timeout information
        driver.set_page_load_timeout(20)
        
        return driver
    
    
    
    def get_element_contents(self, driver, xpath, verbose=True, scroll=True):
        if verbose:
            print('Getting text of {}'.format(xpath))
        results_str = ''
        try:
            if scroll:
                driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
            web_element = self.get_web_element(driver, xpath)
            if scroll:
                driver.execute_script('arguments[0].scrollIntoView(true)', web_element)
            results_str = web_element.text.strip()
        except Exception as e:
            message = str(e).strip()
            if verbose:
                print('Waiting for the web element to be located: {}'.format(message))
            if ('obscures' in message) or ('Other element would receive the click' in message):
                tag_list = obscure_regex.findall(message)
                while len(tag_list) == 2:
                    unobscure_element(driver, tag_list[1])
                    try:
                        if scroll:
                            driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
                        web_element = self.get_web_element(driver, xpath)
                        if scroll:
                            driver.execute_script('arguments[0].scrollIntoView(true)', web_element)
                        results_str = web_element.text.strip()
                        message = ''
                    except Exception as e:
                        message = str(e).strip()
                        if verbose:
                            print('Waiting for the web element to be located (unobscured): {}'.format(message))
                    tag_list = obscure_regex.findall(message)
        
        return results_str
    
    
    
    def get_last_sunday(self, today=datetime.now()):
        start = today - timedelta((today.weekday() + 1) % 7)
        
        return start + relativedelta.relativedelta(weekday=relativedelta.SU(-1))
    
    
    
    def get_web_element(self, driver, xpath):
        try:
            web_element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, xpath))
                )
            
            return web_element
            
        except Exception as e:
            message = str(e).strip()
            print('Waiting for the web_element to show up: {}'.format(message))
    
    
    
    def key_in_search(self, driver, input_xpath, field_value, verbose=True):
        if verbose:
            print('Searching for: {}'.format(field_value))
        input_tag = self.get_web_element(driver, input_xpath)
        input_tag.click()
        ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
        input_tag.send_keys(field_value)
        ActionChains(driver).key_down(Keys.ENTER).perform()
    
    
    
    def move_to(self, driver, xpath, verbose=True):
        if verbose:
            print('Moving to {}'.format(xpath))
        try:
            web_element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, xpath))
                )
            ActionChains(driver).move_to_element(web_element).perform()
        except Exception as e:
            message = str(e).strip()
            if verbose:
                print('Waiting for the web element to be visible: {}'.format(message))
            if 'obscures' in message:
                match_obj = obscure_regex.search(message)
                while match_obj:
                    unobscure_element(driver, match_obj)
                    try:
                        web_element = WebDriverWait(driver, 5).until(
                            EC.visibility_of_element_located((By.XPATH, xpath))
                            )
                        ActionChains(driver).move_to_element(web_element).perform()
                        message = ''
                    except Exception as e:
                        message = str(e).strip()
                        if verbose:
                            print('Waiting for the web element to be visible (unobscured): {}'.format(message))
                    match_obj = obscure_regex.search(message)
    
    
    
    def similar(self, a, b):
        
        return SequenceMatcher(None, str(a), str(b)).ratio()
    
    
    
    def unobscure_element(self, driver, match_tuple):
        tag_name = match_tuple[0]
        class_str = '.'.join([cn for cn in match_tuple[1].split(' ') if cn != 'hidden'])
        overlay_css = '{}.{}'.format(tag_name, class_str)
        try:
            
            # Wait for overlay div to show up
            overlay_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, overlay_css))
                )
            print('Hiding <{}'.format(overlay_tag.get_attribute('outerHTML').split('<')[1]))
            driver.execute_script("arguments[0].setAttribute('style','display:none;');", overlay_tag)
        
        except Exception as e:
            message = str(e).strip()
            print('Waiting for {} {} to show up before hiding it: {}'.format(class_str, tag_name, message))
            driver.refresh()
    
    
    
    def wait_for(self, wait_count, verbose=True):
        if verbose:
            print('Waiting for {} seconds'.format(wait_count))
        time.sleep(wait_count)
    
    
    
    def click_the_audio_apply_button(self, driver):
        apply_css = 'input[tabindex="117"]'
        apply_css = 'div.controls-block > div.button-holder > div.button-block > input'
        apply_xpath = '/html/body/div[28]/div/div[3]/div[3]/div[2]/input'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui flyout light buttons-right dialog-editor-block-audio yui3-dd-draggable visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div", "div[3]", "div[3]", "div[2]", "input"]
        apply_xpath = "/{}".format('/'.join(xpath_list))
        apply_tag = self.get_web_element(driver, apply_xpath)
        apply_tag.click()
    
    
    
    def add_an_episode_note(self, driver, note_str):
        self.click_the_podcasting_tab(driver)
        note_css = 'div.field-rhs > div.field-rte-wrapper > div > div.rte.ProseMirror > p'
        note_xpath = '/html/body/div[27]/div/div[2]/div[2]/div[5]/div/div[2]/div[1]/div/div[2]'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui flyout light buttons-right dialog-editor-block-audio yui3-dd-draggable visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div", "div[2]", "div[2]", "div[5]", "div", "div[2]", "div[1]", "div", "div[2]"]
        note_xpath = "/{}".format('/'.join(xpath_list))
        print('Clicking the "Episode Note" text area and pasting in "{}"'.format(note_str))
        note_tag = self.get_web_element(driver, note_xpath)
        ActionChains(driver).move_to_element(note_tag).perform()
        note_tag.click()
        ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
        ActionChains(driver).key_down(Keys.CONTROL).send_keys('x').key_up(Keys.CONTROL).perform()
        note_tag.send_keys(note_str)
    
    
    
    def click_the_edit_overlay(self, driver):
        div_class_str = 'yui3-widget sqs-widget sqs-data-widget sqs-dialog-field sqs-button-vanilla normal-button data-state-loaded'
        edit_css = 'div[class="{}"][data-test="button"][title="Edit"]'.format(div_class_str)
        edit_xpath = '/html/body/div[28]/div[1]/div[2]/div[1]/div[2]/div[1]/div/div[1]/div/div[1]/div/div[2]/div[3]/div/div[2]'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui standard light buttons-right dialog-text-post-editor visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div[1]", "div[2]", "div[1]", "div[2]", "div[1]", "div", "div[1]", "div",
                      "div[1]", "div", "div[2]", "div[3]", "div", "div[2]"]
        edit_xpath = "/{}".format('/'.join(xpath_list))
        try:
            print('Clicking the "EDIT" overlay')
            edit_tag = WebDriverWait(driver, 10).until(
                #EC.presence_of_element_located((By.XPATH, edit_xpath))
                EC.presence_of_element_located((By.CSS_SELECTOR, edit_css))
                )
            ActionChains(driver).move_to_element(edit_tag).click().perform()
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the "EDIT" overlay to show up: {}'.format(message))
    
    
    def click_the_podcasting_tab(self, driver):
        tab_css = 'div[data-test="tab-1"]'
        tab_xpath = '/html/body/div[27]/div/div[1]/div[2]/div/div/div/div[2]'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui flyout light buttons-right dialog-editor-block-audio yui3-dd-draggable visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div", "div[1]", "div[2]", "div", "div", "div", "div[2]"]
        tab_xpath = "/{}".format('/'.join(xpath_list))
        print('Clicking the "Podcasting" tab')
        tab_tag = self.get_web_element(driver, tab_xpath)
        tab_tag.click()
    
    
    
    def click_the_save_post_button(self, driver):
        save_css = '.saveAndClose'
        save_css_list = ['div.controls-block', 'div.button-holder', 'div:nth-child(4)',
                         'input[class="saveAndClose"][tabindex="105"][type="button"][data-test="dialog-saveAndClose"][value="Save"]']
        save_css = ' > '.join(save_css_list)
        save_xpath = '/html/body/div[58]/div[1]/div[3]/div[3]/div[4]/input'
        save_xpath = '/html/body/div[85]/div[1]/div[3]/div[3]/div[4]/input'
        save_xpath = '/html/body/div[15]/div[1]/div[3]/div[3]/div[4]/input'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui standard light buttons-right dialog-text-post-editor visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div[1]", "div[3]", "div[3]", "div[4]", "input"]
        save_xpath = "/{}".format('/'.join(xpath_list))
        print('Clicking the "SAVE" button')
        save_tag = self.get_web_element(driver, save_xpath)
        save_tag.click()
    
    
    
    def click_the_upload_file_subtab(self, driver):
        tab_css = 'div[class="option-title"]'
        tab_xpath = '/html/body/div[28]/div/div[2]/div[1]/div[2]/div/div[1]/div/div'
        div_class_str = 'standard-dialog-wrapper squarespace-managed-ui flyout light buttons-right dialog-editor-block-audio yui3-dd-draggable visible'
        xpath_list = ["html", "body", self.get_div(div_class_str), "div", "div[2]", "div[1]", "div[2]", "div", "div[1]", "div", "div"]
        tab_xpath = "/{}".format('/'.join(xpath_list))
        print('Clicking the "Upload File" subtab')
        tab_tag = self.get_web_element(driver, tab_xpath)
        tab_tag.click()
    
    
    
    def click_external_file_subtab(self, driver):
        print('Clicking the external file tab')
        button_css = 'div[data-value="external"]'
        try:
            
            # Wait for button to show up
            button_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, button_css))
            )
            button_tag.click()
            
        except Exception as e:
            raise Exception('Waiting for the external file tab to show up: {}'.format(e))
    
    
    
    def click_the_close_signin_window_link(self, driver):
        print('Clicking the "Click here" link')
        #clickhere_css = 'body > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(1) > p:nth-child(2) > font:nth-child(1)
        #> a:nth-child(2)'
        #clickhere_xpath = '/html/body/table/tbody/tr[2]/td/p/font/a/b'
        clickhere_css = 'a[href="javascript:void(close())"]'
        try:
            
            # Wait for signin button to show up
            clickhere_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, clickhere_css))
            )
            clickhere_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the "Click here" link to show up: {}'.format(message))
    
    
    
    def click_the_member_signin_button(self, driver):
        print('Clicking the "Member Sign In" button')
        #signin_css = 'body > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(1) > form:nth-child(2) > table:nth-child(1)
        #> tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1) > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(4) > td:nth-child(2)
        #> input:nth-child(7)'
        #signin_xpath = '/html/body/table[1]/tbody/tr/td/table/tbody/tr/td[4]/font/a'
        signin_css = 'input[value="Member Sign In"]'
        try:
            
            # Wait for signin button to show up
            signin_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, signin_css))
                )
            signin_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for signin button to show up: {}'.format(message))
    
    
    
    def click_the_signin_button(self, driver):
        #signin_css = '.navleft3e'
        #signin_xpath = '/html/body/table[1]/tbody/tr/td/table/tbody/tr/td[4]/font/a'
        print('Clicking the "Sign in or signup" link')
        signin_css = 'a[href="{}"]'.format(self.signin_url)
        try:
            main_window_handle = None
            while not main_window_handle:
                main_window_handle = driver.current_window_handle
            
            # Wait for signin link to show up
            signin_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, signin_css))
                )
            signin_tag.click()
            signin_window_handle = None
            while not signin_window_handle:
                for handle in driver.window_handles:
                    if handle != main_window_handle:
                        signin_window_handle = handle
                        break
            driver.switch_to.window(signin_window_handle)
            username = 'redeemer-ma-pca'
            self.fill_in_field(driver, 'Source', username)
            password = '191SudburyRoad'
            self.fill_in_field(driver, 'Password', password)
            self.click_the_member_signin_button(driver)
            self.click_the_close_signin_window_link(driver)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for signin link to show up: {}'.format(message))
        
        return main_window_handle



    def click_the_copyright_checkbox(self, driver, verbose=True):
        if verbose:
            print('Clicking the copyright checkbox')
        checkbox_css = 'input[name="copyright"]'
        checkbox_css = '#copyright'
        checkbox_xpath = '/html/body/table[2]/tbody/tr/td/table/tbody/tr[3]/td[2]/div/div/div/div/table/tbody/tr/td/table/tbody/tr/td/form/table[2]/tbody/tr[13]/td[2]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/input'
        try:
            
            # Wait for checkbox to show up
            self.move_to(driver, checkbox_xpath, verbose=verbose)
            checkbox_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, checkbox_css))
                )
            checkbox_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the checkbox to show up: {}'.format(message))



    def fill_in_textarea(self, driver, field_name, field_value):
        print('Filling in the {} field with {}'.format(field_name, field_value))
        try:
            input_css = 'textarea[name="{}"]'.format(field_name)
            
            # Wait for textarea field to show up
            input_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, input_css))
                )
            input_tag.click()
            ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            input_tag.send_keys(field_value)
        
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for textarea field to show up: {}'.format(message))



    def click_the_upload_media_button(self, driver, main_window_handle):
        print('Clicking the upload media button')
        upload_css = 'font.ar2 > a.navleftblack5b'
        try:
            driver.switch_to.window(main_window_handle)
            
            # Wait for button to show up
            upload_tag = WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, upload_css))
                )
            upload_tag.click()
            upload_window_handle = None
            while not upload_window_handle:
                for handle in driver.window_handles:
                    if handle != main_window_handle:
                        upload_window_handle = handle
                        break
            driver.switch_to.window(upload_window_handle)
            for frame_tag in driver.find_elements_by_tag_name('frame'):
                if frame_tag.get_attribute('name') == 'AudioAdd1':
                    driver.switch_to.frame(frame_tag)
                    break
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the ID to show up: {}'.format(message))
        
        return upload_window_handle



    def fill_in_the_password_field(self, driver, upload_window_handle):
        print('Filling in the password field')
        password_css = '#password'
        password_xpath = '//*[@id="password"]'
        try:
            
            # Wait for input to show up
            driver.switch_to.window(upload_window_handle)
            driver.maximize_window()
            for frame_tag in driver.find_elements_by_tag_name('frame'):
                if frame_tag.get_attribute('name') == 'AudioAdd1':
                    driver.switch_to.frame(frame_tag)
                    break
            password_tag = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, password_xpath))
                )
            password_tag.click()
            ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            password_tag.send_keys('Genesis11')
            
        except Exception as e:
            message = str(e).strip()
            print('Waiting for the password field to show up (first time): {}'.format(message))
            try:
                
                # Wait for input to show up
                driver.switch_to.window(upload_window_handle)
                driver.maximize_window()
                for frame_tag in driver.find_elements_by_tag_name('frame'):
                    if frame_tag.get_attribute('name') == 'AudioAdd1':
                        driver.switch_to.frame(frame_tag)
                        break
                password_tag = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, password_css))
                    )
                password_tag.click()
                ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
                password_tag.send_keys('Genesis11')
                
            except Exception as e:
                message = str(e).strip()
                raise Exception('Waiting for the password field to show up (second time): {}'.format(message))



    def click_the_browse_button(self, driver, file_path):
        print('Clicking the browse button')
        browse_css = 'input[name="datafile"]'
        try:
            
            # Wait for button to show up
            browse_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, browse_css))
            )
            browse_tag.send_keys(file_path)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the browse window to show up: {}'.format(message))
    
    
    
    def click_another_upload_media_button(self, driver):
        print('Clicking the secondary upload media button')
        upload_css = 'input[name="action_upload"]'
        try:
            
            # Wait for button to show up
            upload_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, upload_css))
                )
            upload_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the button to show up: {}'.format(message))



    def click_the_close_upload_window_link(self, driver, main_window_handle):
        print('Uploading the MP3 and waiting for processing...')
        upload_window_handle = None
        while not upload_window_handle:
            for handle in driver.window_handles:
                if handle != main_window_handle:
                    upload_window_handle = handle
                    break
        driver.switch_to.window(upload_window_handle)
        for frame_tag in driver.find_elements_by_tag_name('frame'):
            if frame_tag.get_attribute('name') == 'AudioAdd1':
                driver.switch_to.frame(frame_tag)
                break
        close_css = 'body > table > tbody > tr:nth-child(2) > td > font > p:nth-child(6) > a'
        try:
            close_tag = WebDriverWait(driver, 140).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, close_css))
                )
            close_tag.click()
            driver.switch_to.window(main_window_handle)
        except Exception as e:
            message = str(e).strip()
            print('Waiting for the close window link to show up (first attempt): {}'.format(message))
            close_xpath = '/html/body/table/tbody/tr[2]/td/font/p[3]/a'
            try:
                close_tag = WebDriverWait(driver, 140).until(
                    EC.presence_of_element_located((By.XPATH, close_xpath))
                    )
                close_tag.click()
                driver.switch_to.window(main_window_handle)
            except Exception as e:
                message = str(e).strip()
                print('Waiting for the close window link to show up (second attempt): {}'.format(message))
                try:
                    driver.execute_script('top.close();')
                    driver.switch_to.window(main_window_handle)
                except Exception as e:
                    message = str(e).strip()
                    raise Exception('Executing top.close() javascript (final attempt): {}'.format(message))



    def cancel_upload_window(self, driver, upload_window_handle, upload_url):
        print('Clicking the cancel upload button')
        cancel_css = 'input[name="action_cancel"]'
        try:
            
            # Wait for button to show up
            driver.switch_to.window(upload_window_handle)
            cancel_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, cancel_css))
                )
            cancel_tag.click()
            if driver.current_url == upload_url:
                driver.close()
            
        except Exception as e:
            message = str(e).strip()
            print('Waiting for the cancel upload button to show up: {}'.format(message))



    def click_the_edit_media_button(self, driver, edit_url, main_window_handle):
        print('Clicking the edit media button')
        edit_css = 'a[href="{}"'.format(edit_url)
        try:
            
            # Wait for button to show up
            edit_tag = WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, edit_css))
                )
            edit_tag.click()
            edit_window_handle = None
            while not edit_window_handle:
                for handle in driver.window_handles:
                    if handle != main_window_handle:
                        edit_window_handle = handle
                        break
            driver.switch_to.window(edit_window_handle)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the ID to show up: {}'.format(message))
        
        return edit_window_handle



    def click_the_speaker_selector_button(self, driver):
        print('Clicking the speaker selector button')
        button_css = 'a[href="winedit_speaker.asp"]'
        try:
            main_window_handle = None
            while not main_window_handle:
                main_window_handle = driver.current_window_handle
            
            # Wait for button to show up
            button_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, button_css))
                )
            button_tag.click()
            signin_window_handle = None
            while not signin_window_handle:
                for handle in driver.window_handles:
                    if handle != main_window_handle:
                        signin_window_handle = handle
                        break
            driver.switch_to.window(signin_window_handle)
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for speaker selector button to show up: {}'.format(message))
        
        return main_window_handle



    def click_the_add_new_speaker_button(self, driver):
        print('Clicking the add new speaker button')
        button_css = 'input[value="Add New Speaker"]'
        try:
            
            # Wait for button to show up
            button_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, button_css))
                )
            button_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the seearch button to show up: {}'.format(message))



    def click_the_new_speaker_checkbox(self, driver):
        print('Clicking the new speaker checkbox')
        checkbox_css = 'input[name="speakercheck"]'
        try:
            
            # Wait for checkbox to show up
            checkbox_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, checkbox_css))
                )
            checkbox_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the new speaker checkbox to show up: {}'.format(message))



    def click_the_first_link(self, driver, search_str):
        print('Clicking the first link')
        link_css = 'a.addtocartlink:nth-child(1)'
        try:
            
            # Wait for link to show up
            link_css = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, link_css))
                )
            link_css.click()
            
        except Exception as e:
            message = str(e).strip()
            print('Adding a new speaker ({}): {}'.format(search_str, message))
            self.click_the_new_speaker_checkbox(driver)
            self.fill_in_field(driver, 'AddSpeaker', search_str)
            self.click_the_add_new_speaker_button(driver)



    def click_the_search_button(self, driver):
        print('Clicking the search button')
        button_css = 'input[name="search"]'
        try:
            
            # Wait for button to show up
            button_tag = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, button_css))
            )
            button_tag.click()
            
        except Exception as e:
            message = str(e).strip()
            raise Exception('Waiting for the seearch button to show up: {}'.format(message))



    def get_mp3_filepath(self, mp3_url):
        parts_list = mp3_url.split('/')
        file_name = urllib.parse.quote(parts_list[-1])
        url_prefix = parts_list[:-1]
        mp3_url = '/'.join(url_prefix+[file_name])
        print('Getting MP3 file path for {}'.format(mp3_url))
        mp3_dir = '../data/mp3'
        os.makedirs(name=mp3_dir, exist_ok=True)
        file_path = os.path.abspath(os.path.join(mp3_dir, file_name))
        if not Path(file_path).is_file():
            path_tuple = urlretrieve(url=mp3_url, filename=file_path)
            file_path = path_tuple[0]
        
        return file_path



    def get_speaker_str(self, id_num, library_items_div):
        print('Getting the speaker')
        speaker_selector = '#audioInfo{} > div > div:nth-child(2) > div:nth-child(1)'.format(id_num)
        speaker_str = library_items_div.select(speaker_selector)[0].text.strip()
        speaker_regex = re.compile(r'(Mr\. |Pasor |Pastor |Rev\. |pastor )?([^\r\n]+)')
        speaker_str = speaker_regex.sub(r'\g<2>', speaker_str)
        if speaker_str == 'Kerr':
            speaker_str = 'Matthew Kerr'
        if speaker_str == 'Andrew Davis':
            speaker_str = 'Andrew Franklin Davis'
        
        return speaker_str
